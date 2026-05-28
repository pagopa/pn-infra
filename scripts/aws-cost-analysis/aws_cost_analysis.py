#!/usr/bin/env python3
"""
AWS Cost Analysis Script

Fetches daily costs from AWS Cost Explorer grouped by Service and Tag "Microservice"
for the last 6 months across multiple environments and AWS accounts, then generates:
  - One Excel workbook (.xlsx) per environment
  - A single interactive HTML dashboard with filters for env, account, service, tag, month

Environments : dev, uat, test, hotfix, prod
SSO profiles  : sso_pn-core-<env>  and  sso_pn-confinfo-<env>

Usage:
    # Single / multiple environments
    python aws_cost_analysis.py --env dev prod

    # All environments
    python aws_cost_analysis.py --all

    # Custom output directory
    python aws_cost_analysis.py --all --output-dir ./reports

Requirements:
    pip install -r requirements.txt
"""

import argparse
import json
import sys
from datetime import date
from pathlib import Path

import boto3
import pandas as pd
from botocore.exceptions import ClientError, NoCredentialsError, TokenRetrievalError
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

VALID_ENVS   = ["dev", "uat", "test", "hotfix", "prod"]
ACCOUNT_TYPES = ["core", "confinfo"]

HEADER_FILL    = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT    = Font(color="FFFFFF", bold=True)
ALT_ROW_FILL   = PatternFill("solid", fgColor="D6E4F0")
TREND_UP_FILL  = PatternFill("solid", fgColor="FF9999")
TREND_DOWN_FILL = PatternFill("solid", fgColor="AAFFAA")


def _date_range_6_months() -> tuple[str, str]:
    """Return (start_date, end_date) covering the last 6 complete months up to today."""
    today = date.today()
    end = today.replace(day=1)
    month = end.month - 6
    year = end.year
    if month <= 0:
        month += 12
        year -= 1
    start = date(year, month, 1)
    return start.isoformat(), end.isoformat()


def _month_label(iso_date: str) -> str:
    """'2024-03-15' -> '2024-03'"""
    return iso_date[:7]


def _profile_name(env: str, account_type: str) -> str:
    return f"sso_pn-{account_type}-{env}"


# ---------------------------------------------------------------------------
# AWS data fetching
# ---------------------------------------------------------------------------

def fetch_costs(profile: str, env: str, account_type: str) -> list[dict]:
    """
    Query Cost Explorer for daily costs grouped by SERVICE and tag 'Microservice'.
    Returns records: {date, service, microservice, amount, env, account}.
    Returns [] and prints a warning on auth errors (non-fatal, allows other accounts to proceed).
    """
    try:
        session = boto3.Session(profile_name=profile)
        ce = session.client("ce", region_name="us-east-1")
    except Exception as exc:
        print(f"  [WARN] Cannot create session for profile '{profile}': {exc}")
        return []

    start, end = _date_range_6_months()
    print(f"  [{env}/{account_type}] Fetching {start} → {end} (profile: {profile}) ...")

    records: list[dict] = []
    next_token: str | None = None

    while True:
        kwargs: dict = dict(
            TimePeriod={"Start": start, "End": end},
            Granularity="DAILY",
            Metrics=["UnblendedCost"],
            GroupBy=[
                {"Type": "DIMENSION", "Key": "SERVICE"},
                {"Type": "TAG",       "Key": "Microservice"},
            ],
        )
        if next_token:
            kwargs["NextPageToken"] = next_token

        try:
            response = ce.get_cost_and_usage(**kwargs)
        except (NoCredentialsError, TokenRetrievalError) as exc:
            print(f"  [WARN] Auth error for profile '{profile}': {exc}")
            print(f"         Run: aws sso login --profile {profile}")
            return []
        except ClientError as exc:
            print(f"  [WARN] ClientError for profile '{profile}': {exc}")
            return []

        for result_by_time in response["ResultsByTime"]:
            day = result_by_time["TimePeriod"]["Start"]
            for group in result_by_time["Groups"]:
                service = group["Keys"][0]
                if service == "Tax":
                    continue
                tag_raw = group["Keys"][1]  # e.g. "Microservice$pn-delivery"
                microservice = tag_raw.split("$", 1)[-1] if "$" in tag_raw else ""
                amount = float(group["Metrics"]["UnblendedCost"]["Amount"])
                if amount > 0:
                    records.append(
                        {
                            "date": day,
                            "env": env,
                            "account": account_type,
                            "service": service,
                            "microservice": microservice if microservice else "(untagged)",
                            "amount": round(amount, 6),
                        }
                    )

        next_token = response.get("NextPageToken")
        if not next_token:
            break

    print(f"  [{env}/{account_type}] → {len(records)} records.")
    return records


# ---------------------------------------------------------------------------
# Data processing
# ---------------------------------------------------------------------------

def build_dataframes(records: list[dict]) -> dict:
    """Return processed DataFrames from a list of cost records (any subset)."""
    df = pd.DataFrame(records)
    if df.empty:
        return {}

    df["month"] = df["date"].apply(_month_label)
    months = sorted(df["month"].unique())

    def _pivot(group_col: str) -> pd.DataFrame:
        pv = (
            df.groupby([group_col, "month"])["amount"]
            .sum().unstack(fill_value=0).round(2)
        )
        pv = pv.reindex(columns=months, fill_value=0)
        pv["Total"] = pv.sum(axis=1)
        return pv.sort_values("Total", ascending=False)

    by_service = _pivot("service")
    by_tag      = _pivot("microservice")
    by_account  = _pivot("account")

    daily = df.sort_values(["date", "env", "account", "service", "microservice"])

    # Month-over-month delta for the two most recent months
    if len(months) >= 2:
        last, prev = months[-1], months[-2]

        def _trend(pv: pd.DataFrame, row_col: str) -> pd.DataFrame:
            t = pd.DataFrame({
                row_col:     pv.index,
                prev:        pv[prev].values,
                last:        pv[last].values,
            })
            t["delta_abs"] = (t[last] - t[prev]).round(2)
            t["delta_pct"] = (
                (t["delta_abs"] / t[prev].replace(0, float("nan"))) * 100
            ).round(1)
            return t.sort_values("delta_abs", key=abs, ascending=False)

        trend_service = _trend(by_service, "service")
        trend_tag     = _trend(by_tag,     "microservice")
    else:
        trend_service = pd.DataFrame()
        trend_tag     = pd.DataFrame()

    return {
        "by_service":    by_service,
        "by_tag":        by_tag,
        "by_account":    by_account,
        "daily":         daily,
        "trend_service": trend_service,
        "trend_tag":     trend_tag,
        "months":        months,
    }


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def _write_pivot_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, row_label: str):
    ws = wb.create_sheet(title=sheet_name)
    # Header
    headers = [row_label] + list(df.columns)
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Rows
    for row_num, (idx, row) in enumerate(df.iterrows(), start=2):
        fill = ALT_ROW_FILL if row_num % 2 == 0 else None
        ws.cell(row=row_num, column=1, value=idx)
        for col_idx, val in enumerate(row, start=2):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            if fill:
                cell.fill = fill
            if isinstance(val, float):
                cell.number_format = '#,##0.00'

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Simple bar chart for top-10 totals (last column = Total)
    if "Total" in df.columns:
        top10 = min(10, len(df))
        chart = BarChart()
        chart.type = "bar"
        chart.title = f"Top {top10} by Total Cost (USD)"
        chart.y_axis.title = row_label
        chart.x_axis.title = "USD"
        chart.width = 25
        chart.height = 14
        total_col = len(headers)
        data = Reference(ws, min_col=total_col, max_col=total_col, min_row=1, max_row=top10 + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=top10 + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{len(df) + 4}")


def _write_trend_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, row_label: str):
    if df.empty:
        return
    ws = wb.create_sheet(title=sheet_name)
    headers = list(df.columns)
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_num, (_, row) in enumerate(df.iterrows(), start=2):
        delta = row.get("delta_abs", 0)
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            if isinstance(val, float):
                cell.number_format = '#,##0.00'
            if col_idx >= 4:  # delta columns
                if delta > 0:
                    cell.fill = TREND_UP_FILL
                elif delta < 0:
                    cell.fill = TREND_DOWN_FILL

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def _write_daily_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet(title="Daily Detail")
    has_multi = "env" in df.columns
    headers = ["Date", "Env", "Account", "Service", "Microservice", "Amount (USD)"] if has_multi \
              else ["Date", "Service", "Microservice", "Amount (USD)"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for row_num, (_, row) in enumerate(df.iterrows(), start=2):
        fill = ALT_ROW_FILL if row_num % 2 == 0 else None
        vals = ([row["date"], row.get("env", ""), row.get("account", ""),
                 row["service"], row["microservice"], row["amount"]]
                if has_multi else
                [row["date"], row["service"], row["microservice"], row["amount"]])
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            if fill:
                cell.fill = fill
            if col_idx == len(headers):
                cell.number_format = '#,##0.000000'

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def generate_xlsx(dfs: dict, output_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    _write_pivot_sheet(wb, "By Service",        dfs["by_service"], "Service")
    _write_pivot_sheet(wb, "By Microservice Tag", dfs["by_tag"],   "Microservice")
    _write_pivot_sheet(wb, "By Account",        dfs["by_account"], "Account")
    _write_trend_sheet(wb, "Trend by Service",  dfs["trend_service"], "service")
    _write_trend_sheet(wb, "Trend by Tag",      dfs["trend_tag"],     "microservice")
    _write_daily_sheet(wb, dfs["daily"])

    wb.save(output_path)
    print(f"  → Excel saved to: {output_path}")


# ---------------------------------------------------------------------------
# HTML dashboard generation
# ---------------------------------------------------------------------------

def _to_js(obj) -> str:
    return json.dumps(obj, default=str)


def generate_html(all_records: list[dict], output_path: str):
    """
    Generate a single self-contained HTML dashboard from records that span
    multiple environments and accounts.  All filtering and trend computation
    happens client-side in JS so every filter combination is reactive.
    """
    df = pd.DataFrame(all_records)
    months   = sorted(df["date"].apply(_month_label).unique().tolist())
    envs     = sorted(df["env"].unique().tolist())
    accounts = sorted(df["account"].unique().tolist())
    services = sorted(df["service"].unique().tolist())
    tags     = sorted(df["microservice"].unique().tolist())

    records_json = _to_js(
        df[["date", "env", "account", "service", "microservice", "amount"]]
        .to_dict("records")
    )

    date_from = months[0] if months else ""
    date_to   = months[-1] if months else ""

    palette = [
        "#2196F3","#4CAF50","#F44336","#FF9800","#9C27B0","#00BCD4",
        "#8BC34A","#E91E63","#3F51B5","#009688","#FFC107","#795548",
        "#607D8B","#CDDC39","#FF5722","#673AB7","#03A9F4","#76FF03",
    ]

    env_opts     = "\n".join(f'        <option value="{e}">{e}</option>' for e in envs)
    account_opts = "\n".join(f'        <option value="{a}">{a}</option>' for a in accounts)
    service_opts = "\n".join(f'        <option value="{s}">{s}</option>' for s in services)
    tag_opts     = "\n".join(f'        <option value="{t}">{t}</option>' for t in tags)
    month_opts   = "\n".join(f'        <option value="{m}">{m}</option>' for m in months)

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>AWS Cost Analysis – Last 6 Months</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.2/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-datalabels@2.2.0/dist/chartjs-plugin-datalabels.min.js"></script>
<style>
  :root {{
    --primary:#1F4E79; --accent:#2196F3; --bg:#F0F4F8;
    --card:#FFFFFF; --text:#1A1A2E; --muted:#6B7280;
  }}
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{font-family:'Segoe UI',system-ui,sans-serif;background:var(--bg);color:var(--text);}}
  header{{background:var(--primary);color:#fff;padding:1.2rem 2rem;display:flex;align-items:center;gap:1rem;}}
  header h1{{font-size:1.4rem;}}
  header span{{font-size:.85rem;opacity:.7;}}
  .container{{max-width:1500px;margin:0 auto;padding:1.5rem 1rem;}}
  .cards{{display:flex;flex-wrap:wrap;gap:1rem;margin-bottom:1.5rem;}}
  .card{{background:var(--card);border-radius:10px;padding:1rem 1.4rem;
         box-shadow:0 1px 4px rgba(0,0,0,.08);min-width:150px;flex:1;}}
  .card .label{{font-size:.72rem;color:var(--muted);text-transform:uppercase;letter-spacing:.05em;}}
  .card .value{{font-size:1.45rem;font-weight:700;color:var(--primary);}}
  .filters{{background:var(--card);border-radius:10px;padding:1rem 1.4rem;
            box-shadow:0 1px 4px rgba(0,0,0,.08);margin-bottom:1.5rem;
            display:flex;flex-wrap:wrap;gap:1rem;align-items:flex-end;}}
  .filter-group{{display:flex;flex-direction:column;gap:.3rem;min-width:170px;flex:1;}}
  .filter-group label{{font-size:.8rem;font-weight:600;color:var(--muted);}}
  select{{border:1px solid #D1D5DB;border-radius:6px;padding:.45rem .7rem;
          font-size:.88rem;background:#fff;width:100%;}}
  select[multiple]{{height:110px;}}
  button{{padding:.5rem 1.2rem;border:none;border-radius:6px;background:var(--accent);
          color:#fff;font-size:.9rem;cursor:pointer;font-weight:600;transition:opacity .2s;}}
  button:hover{{opacity:.85;}}
  button.secondary{{background:#6B7280;}}
  .grid2{{display:grid;grid-template-columns:1fr 1fr;gap:1.5rem;margin-bottom:1.5rem;}}
  .grid3{{display:grid;grid-template-columns:1fr 1fr 1fr;gap:1.5rem;margin-bottom:1.5rem;}}
  @media(max-width:1000px){{.grid2,.grid3{{grid-template-columns:1fr;}}}}
  .panel{{background:var(--card);border-radius:10px;padding:1.2rem;
          box-shadow:0 1px 4px rgba(0,0,0,.08);}}
  .panel h2{{font-size:.95rem;margin-bottom:.9rem;color:var(--primary);}}
  .full-panel{{grid-column:1/-1;}}
  .chart-wrap{{position:relative;height:300px;}}
  .chart-wrap-md{{position:relative;height:360px;}}
  .table-wrap{{overflow-x:auto;max-height:380px;overflow-y:auto;}}
  table{{width:100%;border-collapse:collapse;font-size:.83rem;}}
  thead th{{position:sticky;top:0;z-index:1;background:var(--primary);color:#fff;
             padding:.5rem .8rem;text-align:left;white-space:nowrap;}}
  th.sortable{{cursor:pointer;user-select:none;}}
  th.sortable:hover{{background:#2563EB;}}
  td{{padding:.42rem .8rem;border-bottom:1px solid #E5E7EB;}}
  tr:hover td{{background:#EFF6FF;}}
  tr:nth-child(even) td{{background:#F8FAFC;}}
  .badge{{display:inline-block;padding:.1rem .45rem;border-radius:99px;font-size:.72rem;font-weight:600;}}
  .up{{background:#FEE2E2;color:#B91C1C;}}
  .down{{background:#DCFCE7;color:#15803D;}}
  .neutral{{background:#F3F4F6;color:#374151;}}
  .env-pill{{display:inline-block;padding:.1rem .5rem;border-radius:4px;font-size:.75rem;font-weight:600;background:#DBEAFE;color:#1E40AF;}}
  .acct-pill{{display:inline-block;padding:.1rem .5rem;border-radius:4px;font-size:.75rem;font-weight:600;background:#EDE9FE;color:#5B21B6;}}
  .section-title{{font-size:1.05rem;font-weight:700;color:var(--primary);margin:1.5rem 0 .8rem;}}
  .pagination{{display:flex;gap:.4rem;align-items:center;margin-top:.8rem;flex-wrap:wrap;}}
  .page-btn{{padding:.22rem .55rem;border:1px solid #D1D5DB;border-radius:4px;
              background:#fff;cursor:pointer;font-size:.78rem;}}
  .page-btn.active{{background:var(--primary);color:#fff;border-color:var(--primary);}}
  #page-info{{font-size:.78rem;color:var(--muted);margin-top:.4rem;}}
</style>
</head>
<body>
<header>
  <div>
    <h1>☁️ AWS Cost Analysis</h1>
    <span>Last 6 months &nbsp;·&nbsp; {date_from} → {date_to} &nbsp;·&nbsp;
          Envs: {", ".join(envs)} &nbsp;·&nbsp; Grouped by Service &amp; Microservice tag</span>
  </div>
</header>
<div class="container">

  <!-- Summary cards -->
  <div class="cards" id="summary-cards"></div>

  <!-- Filters -->
  <div class="filters">
    <div class="filter-group">
      <label>Environment</label>
      <select id="f-env" multiple>
        <option value="__ALL__" selected>All Environments</option>
{env_opts}
      </select>
    </div>
    <div class="filter-group">
      <label>Account</label>
      <select id="f-account" multiple>
        <option value="__ALL__" selected>All Accounts</option>
{account_opts}
      </select>
    </div>
    <div class="filter-group">
      <label>Service</label>
      <select id="f-service" multiple>
        <option value="__ALL__" selected>All Services</option>
{service_opts}
      </select>
    </div>
    <div class="filter-group">
      <label>Microservice Tag</label>
      <select id="f-tag" multiple>
        <option value="__ALL__" selected>All Tags</option>
{tag_opts}
      </select>
    </div>
    <div class="filter-group" style="max-width:150px;">
      <label>Month</label>
      <select id="f-month">
        <option value="__ALL__">All Months</option>
{month_opts}
      </select>
    </div>
    <div style="display:flex;gap:.5rem;align-items:flex-end;">
      <button onclick="applyFilters()">Apply</button>
      <button class="secondary" onclick="resetFilters()">Reset</button>
    </div>
  </div>

  <!-- Row 1: Monthly total (full-width stacked by env) -->
  <div class="grid2">
    <div class="panel full-panel">
      <h2>Monthly Total Cost – stacked by Environment (USD)</h2>
      <div class="chart-wrap"><canvas id="chart-monthly-env"></canvas></div>
    </div>
  </div>

  <!-- Row 2: by service / by tag -->
  <div class="grid2">
    <div class="panel">
      <h2>Cost by Service – Top 15 (USD)</h2>
      <div class="chart-wrap-md"><canvas id="chart-service"></canvas></div>
    </div>
    <div class="panel">
      <h2>Cost by Microservice Tag – Top 15 (USD)</h2>
      <div class="chart-wrap-md"><canvas id="chart-tag"></canvas></div>
    </div>
  </div>

  <!-- Row 3: by account / daily -->
  <div class="grid2">
    <div class="panel">
      <h2>Cost by Account (USD)</h2>
      <div class="chart-wrap"><canvas id="chart-account"></canvas></div>
    </div>
    <div class="panel">
      <h2>Daily Cost Trend</h2>
      <div class="chart-wrap"><canvas id="chart-daily"></canvas></div>
    </div>
  </div>

  <!-- Trend tables (computed in JS from filtered data) -->
  <div class="section-title">📈 Month-over-Month Trend Highlights <span id="trend-months-label" style="font-size:.85rem;font-weight:400;color:var(--muted);"></span></div>
  <div class="grid2">
    <div class="panel">
      <h2>By Service</h2>
      <div class="table-wrap" id="trend-service-table"></div>
    </div>
    <div class="panel">
      <h2>By Microservice Tag</h2>
      <div class="table-wrap" id="trend-tag-table"></div>
    </div>
  </div>

  <!-- Detail table -->
  <div class="panel">
    <h2>📋 Daily Cost Detail</h2>
    <div class="table-wrap">
      <table>
        <thead>
          <tr>
            <th class="sortable" onclick="sortTable(0)">Date ⇅</th>
            <th class="sortable" onclick="sortTable(1)">Env ⇅</th>
            <th class="sortable" onclick="sortTable(2)">Account ⇅</th>
            <th class="sortable" onclick="sortTable(3)">Service ⇅</th>
            <th class="sortable" onclick="sortTable(4)">Microservice ⇅</th>
            <th class="sortable" onclick="sortTable(5)">Amount (USD) ⇅</th>
          </tr>
        </thead>
        <tbody id="detail-tbody"></tbody>
      </table>
    </div>
    <div class="pagination" id="pagination"></div>
    <div id="page-info"></div>
  </div>

</div><!-- /container -->

<script>
// ── Data ────────────────────────────────────────────────────────────────────
const ALL_RECORDS = {records_json};
const ALL_MONTHS  = {_to_js(months)};
const PALETTE     = {_to_js(palette)};

// ── State ────────────────────────────────────────────────────────────────────
let filtered   = [...ALL_RECORDS];
let currentPage = 1;
const PAGE_SIZE = 100;
let sortCol = 0, sortAsc = true;
let charts  = {{}};

// ── Filter helpers ────────────────────────────────────────────────────────────
function getMulti(id) {{
  const vals = Array.from(document.getElementById(id).selectedOptions).map(o => o.value);
  return vals.includes('__ALL__') || vals.length === 0 ? null : vals;
}}
function getSingle(id) {{
  const v = document.getElementById(id).value;
  return v === '__ALL__' ? null : v;
}}
function resetSelect(id) {{
  const sel = document.getElementById(id);
  for (const opt of sel.options) opt.selected = (opt.value === '__ALL__');
}}

function applyFilters() {{
  const envs    = getMulti('f-env');
  const accounts= getMulti('f-account');
  const svcs    = getMulti('f-service');
  const tags    = getMulti('f-tag');
  const month   = getSingle('f-month');
  filtered = ALL_RECORDS.filter(r => {{
    if (envs     && !envs.includes(r.env))          return false;
    if (accounts && !accounts.includes(r.account))  return false;
    if (svcs     && !svcs.includes(r.service))       return false;
    if (tags     && !tags.includes(r.microservice))  return false;
    if (month    && !r.date.startsWith(month))       return false;
    return true;
  }});
  currentPage = 1;
  renderAll();
}}

function resetFilters() {{
  ['f-env','f-account','f-service','f-tag'].forEach(resetSelect);
  document.getElementById('f-month').value = '__ALL__';
  filtered = [...ALL_RECORDS];
  currentPage = 1;
  renderAll();
}}

// ── Aggregation helpers ───────────────────────────────────────────────────────
function groupSum(records, key) {{
  const m = {{}};
  records.forEach(r => {{ m[r[key]] = (m[r[key]] || 0) + r.amount; }});
  return m;
}}
function groupByMonthAndKey(records, key) {{
  // returns {{ month: {{ keyVal: amount }} }}
  const m = {{}};
  records.forEach(r => {{
    const mo = r.date.slice(0, 7);
    if (!m[mo]) m[mo] = {{}};
    m[mo][r[key]] = (m[mo][r[key]] || 0) + r.amount;
  }});
  return m;
}}
function groupByMonth(records) {{
  const m = {{}};
  records.forEach(r => {{ const mo = r.date.slice(0,7); m[mo] = (m[mo]||0) + r.amount; }});
  return m;
}}
function groupByDay(records) {{
  const m = {{}};
  records.forEach(r => {{ m[r.date] = (m[r.date]||0) + r.amount; }});
  return Object.fromEntries(Object.entries(m).sort());
}}
function topN(obj, n=15) {{
  return Object.entries(obj).sort((a,b)=>b[1]-a[1]).slice(0,n);
}}
function fmt(n) {{ return (+n).toLocaleString('en-US',{{minimumFractionDigits:2,maximumFractionDigits:2}}); }}

// ── Trend computation (JS-side, reacts to filters) ────────────────────────────
function computeTrend(records, groupKey) {{
  const months = [...new Set(records.map(r => r.date.slice(0,7)))].sort();
  if (months.length < 2) return {{ rows:[], prev:'', last:'' }};
  const prev = months[months.length-2];
  const last = months[months.length-1];
  const pm = {{}}, lm = {{}};
  records.forEach(r => {{
    const mo = r.date.slice(0,7);
    if (mo === prev) pm[r[groupKey]] = (pm[r[groupKey]]||0) + r.amount;
    if (mo === last) lm[r[groupKey]] = (lm[r[groupKey]]||0) + r.amount;
  }});
  const keys = new Set([...Object.keys(pm), ...Object.keys(lm)]);
  const rows = [...keys].map(k => {{
    const p = pm[k]||0, l = lm[k]||0, d = l - p;
    return {{ key:k, prev:p, last:l, delta_abs:d, delta_pct: p ? d/p*100 : null }};
  }});
  return {{ rows: rows.sort((a,b) => Math.abs(b.delta_abs)-Math.abs(a.delta_abs)), prev, last }};
}}

// ── Chart helpers ─────────────────────────────────────────────────────────────
function destroyChart(id) {{
  if (charts[id]) {{ charts[id].destroy(); delete charts[id]; }}
}}

function makeHBar(id, labels, data) {{
  destroyChart(id);
  const ctx = document.getElementById(id).getContext('2d');
  charts[id] = new Chart(ctx, {{
    type:'bar',
    data:{{ labels, datasets:[{{ data,
      backgroundColor: labels.map((_,i)=>PALETTE[i%PALETTE.length]), borderRadius:4
    }}] }},
    options:{{
      responsive:true, maintainAspectRatio:false, indexAxis:'y',
      plugins:{{ legend:{{display:false}}, datalabels:{{display:false}} }},
      scales:{{ x:{{ ticks:{{ callback:v=>'$'+fmt(v) }} }} }}
    }}
  }});
}}

function makeStackedBar(id, labels, datasets) {{
  destroyChart(id);
  const ctx = document.getElementById(id).getContext('2d');
  charts[id] = new Chart(ctx, {{
    type:'bar',
    data:{{ labels, datasets }},
    options:{{
      responsive:true, maintainAspectRatio:false,
      plugins:{{ legend:{{position:'top'}}, datalabels:{{display:false}} }},
      scales:{{
        x:{{ stacked:true }},
        y:{{ stacked:true, ticks:{{ callback:v=>'$'+fmt(v) }} }}
      }}
    }}
  }});
}}

function makeLineChart(id, labels, data) {{
  destroyChart(id);
  const ctx = document.getElementById(id).getContext('2d');
  charts[id] = new Chart(ctx, {{
    type:'line',
    data:{{ labels, datasets:[{{ label:'Daily Cost (USD)', data,
      borderColor:'#2196F3', backgroundColor:'rgba(33,150,243,.1)',
      fill:true, tension:.3, pointRadius:2
    }}] }},
    options:{{
      responsive:true, maintainAspectRatio:false,
      plugins:{{ legend:{{display:false}}, datalabels:{{display:false}} }},
      scales:{{ y:{{ ticks:{{ callback:v=>'$'+fmt(v) }} }} }}
    }}
  }});
}}

// ── Render: Summary cards ─────────────────────────────────────────────────────
function renderCards() {{
  const byM    = groupByMonth(filtered);
  const total  = Object.values(byM).reduce((a,b)=>a+b,0);
  const months = Object.keys(byM).sort();
  const lastTwo = months.slice(-2).map(m =>
    `<div class="card"><div class="label">${{m}}</div><div class="value">$${{fmt(byM[m]||0)}}</div></div>`
  ).join('');
  document.getElementById('summary-cards').innerHTML = `
    <div class="card"><div class="label">Total (period)</div><div class="value">$${{fmt(total)}}</div></div>
    ${{lastTwo}}
    <div class="card"><div class="label">Envs</div><div class="value">${{new Set(filtered.map(r=>r.env)).size}}</div></div>
    <div class="card"><div class="label">Accounts</div><div class="value">${{new Set(filtered.map(r=>r.env+'/'+r.account)).size}}</div></div>
    <div class="card"><div class="label">Services</div><div class="value">${{new Set(filtered.map(r=>r.service)).size}}</div></div>
    <div class="card"><div class="label">Tags</div><div class="value">${{new Set(filtered.map(r=>r.microservice)).size}}</div></div>
  `;
}}

// ── Render: Charts ────────────────────────────────────────────────────────────
function renderCharts() {{
  // Monthly stacked by env
  const envList = [...new Set(filtered.map(r=>r.env))].sort();
  const months  = ALL_MONTHS.filter(m => filtered.some(r=>r.date.startsWith(m)));
  const stackDs = envList.map((env, i) => {{
    const data = months.map(m => {{
      const s = filtered.filter(r=>r.env===env&&r.date.startsWith(m)).reduce((a,r)=>a+r.amount,0);
      return +s.toFixed(2);
    }});
    return {{ label:env, data, backgroundColor:PALETTE[i%PALETTE.length], borderRadius:3 }};
  }});
  makeStackedBar('chart-monthly-env', months, stackDs);

  // By service
  const bySvc = topN(groupSum(filtered,'service'));
  makeHBar('chart-service', bySvc.map(x=>x[0]), bySvc.map(x=>+x[1].toFixed(2)));

  // By tag
  const byTag = topN(groupSum(filtered,'microservice'));
  makeHBar('chart-tag', byTag.map(x=>x[0]), byTag.map(x=>+x[1].toFixed(2)));

  // By account (stacked by env)
  const acctList = [...new Set(filtered.map(r=>r.account))].sort();
  const acctDs   = envList.map((env, i) => {{
    const data = acctList.map(a => {{
      const s = filtered.filter(r=>r.env===env&&r.account===a).reduce((acc,r)=>acc+r.amount,0);
      return +s.toFixed(2);
    }});
    return {{ label:env, data, backgroundColor:PALETTE[i%PALETTE.length], borderRadius:3 }};
  }});
  destroyChart('chart-account');
  const ctx = document.getElementById('chart-account').getContext('2d');
  charts['chart-account'] = new Chart(ctx, {{
    type:'bar',
    data:{{ labels:acctList, datasets:acctDs }},
    options:{{
      responsive:true, maintainAspectRatio:false,
      plugins:{{ legend:{{position:'top'}}, datalabels:{{display:false}} }},
      scales:{{ x:{{ stacked:true }}, y:{{ stacked:true, ticks:{{ callback:v=>'$'+fmt(v) }} }} }}
    }}
  }});

  // Daily line
  const byDay = groupByDay(filtered);
  makeLineChart('chart-daily', Object.keys(byDay), Object.values(byDay).map(v=>+v.toFixed(2)));
}}

// ── Render: Trend tables ──────────────────────────────────────────────────────
function trendBadge(pct) {{
  if (pct === null || pct === undefined || isNaN(pct)) return '<span class="badge neutral">—</span>';
  const cls   = pct>0?'up':(pct<0?'down':'neutral');
  const arrow = pct>0?'▲':(pct<0?'▼':'–');
  return `<span class="badge ${{cls}}">${{arrow}} ${{Math.abs(pct).toFixed(1)}}%</span>`;
}}

function renderTrendTable(containerId, trend, rowKey) {{
  const el = document.getElementById(containerId);
  if (!trend.rows.length) {{ el.innerHTML='<em style="color:var(--muted)">Not enough data for trend.</em>'; return; }}
  document.getElementById('trend-months-label').textContent = `(${{trend.prev}} → ${{trend.last}})`;
  let html = `<table><thead><tr>
    <th>${{rowKey}}</th><th>${{trend.prev}}</th><th>${{trend.last}}</th><th>Δ USD</th><th>Δ %</th>
  </tr></thead><tbody>`;
  trend.rows.slice(0,20).forEach(r => {{
    const cls = r.delta_abs>0?'up':(r.delta_abs<0?'down':'neutral');
    html += `<tr>
      <td>${{r.key}}</td>
      <td>$${{fmt(r.prev)}}</td>
      <td>$${{fmt(r.last)}}</td>
      <td class="${{cls}}">${{r.delta_abs>=0?'+':''}}${{fmt(r.delta_abs)}}</td>
      <td>${{trendBadge(r.delta_pct)}}</td>
    </tr>`;
  }});
  html += '</tbody></table>';
  el.innerHTML = html;
}}

function renderTrends() {{
  const ts = computeTrend(filtered, 'service');
  const tt = computeTrend(filtered, 'microservice');
  renderTrendTable('trend-service-table', ts, 'Service');
  renderTrendTable('trend-tag-table',     tt, 'Microservice');
}}

// ── Render: Detail table ──────────────────────────────────────────────────────
function sortTable(col) {{
  if (sortCol === col) sortAsc = !sortAsc; else {{ sortCol=col; sortAsc=true; }}
  renderDetailTable();
}}

function renderDetailTable() {{
  const keys = ['date','env','account','service','microservice','amount'];
  const sorted = [...filtered].sort((a,b) => {{
    const av=a[keys[sortCol]], bv=b[keys[sortCol]];
    if (typeof av==='number') return sortAsc?av-bv:bv-av;
    return sortAsc?String(av).localeCompare(String(bv)):String(bv).localeCompare(String(av));
  }});
  const total  = sorted.length;
  const pages  = Math.ceil(total/PAGE_SIZE);
  const start  = (currentPage-1)*PAGE_SIZE;
  const page   = sorted.slice(start, start+PAGE_SIZE);
  document.getElementById('detail-tbody').innerHTML = page.map(r=>`
    <tr>
      <td>${{r.date}}</td>
      <td><span class="env-pill">${{r.env}}</span></td>
      <td><span class="acct-pill">${{r.account}}</span></td>
      <td>${{r.service}}</td>
      <td>${{r.microservice}}</td>
      <td>$${{fmt(r.amount)}}</td>
    </tr>`).join('');
  let pag='';
  for(let p=1;p<=Math.min(pages,20);p++)
    pag+=`<button class="page-btn ${{p===currentPage?'active':''}}" onclick="gotoPage(${{p}})">${{p}}</button>`;
  if(pages>20) pag+='<span style="padding:.2rem">…</span>';
  document.getElementById('pagination').innerHTML=pag;
  document.getElementById('page-info').textContent=
    `Showing ${{start+1}}–${{Math.min(start+PAGE_SIZE,total)}} of ${{total}} records`;
}}
function gotoPage(p){{currentPage=p;renderDetailTable();}}

// ── Main render ───────────────────────────────────────────────────────────────
function renderAll() {{
  renderCards();
  renderCharts();
  renderTrends();
  renderDetailTable();
}}

Chart.register(ChartDataLabels);
renderAll();
</script>
</body>
</html>
"""

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  → HTML dashboard saved to: {output_path}")


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=(
            "AWS Cost Analysis – fetches costs for one or more environments via SSO profiles "
            "(sso_pn-core-<env> and sso_pn-confinfo-<env>), generates one XLSX per environment "
            "and a single interactive HTML dashboard."
        )
    )
    env_group = parser.add_mutually_exclusive_group(required=True)
    env_group.add_argument(
        "--env", nargs="+", choices=VALID_ENVS, metavar="ENV",
        help=f"One or more environments: {', '.join(VALID_ENVS)}",
    )
    env_group.add_argument(
        "--all", action="store_true",
        help="Include all environments",
    )
    parser.add_argument(
        "--output-dir", default="./output",
        help="Directory where output files are saved (default: ./output)",
    )
    args = parser.parse_args()

    envs = VALID_ENVS if args.all else args.env
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    today_str = date.today().strftime("%Y%m%d")

    all_records: list[dict] = []
    xlsx_paths: list[str]   = []

    # ── Fetch per env / account ────────────────────────────────────────────
    for env in envs:
        env_records: list[dict] = []
        for account_type in ACCOUNT_TYPES:
            profile = _profile_name(env, account_type)
            recs = fetch_costs(profile, env, account_type)
            env_records.extend(recs)
            all_records.extend(recs)

        if not env_records:
            print(f"  [WARN] No data for env '{env}' – skipping XLSX.")
            continue

        # ── Per-env XLSX ───────────────────────────────────────────────────
        print(f"  Processing data for env '{env}' ...")
        dfs = build_dataframes(env_records)
        xlsx_path = str(output_dir / f"aws_cost_analysis_{env}_{today_str}.xlsx")
        print(f"  Generating Excel for env '{env}' ...")
        generate_xlsx(dfs, xlsx_path)
        xlsx_paths.append(xlsx_path)

    if not all_records:
        print("\nNo data returned for any environment – exiting.")
        sys.exit(0)

    # ── Combined HTML dashboard ────────────────────────────────────────────
    html_path = str(output_dir / f"aws_cost_analysis_{today_str}.html")
    print("\nGenerating combined HTML dashboard ...")
    generate_html(all_records, html_path)

    # ── Console trend summary ──────────────────────────────────────────────
    print("\n" + "=" * 65)
    print("MONTH-OVER-MONTH TREND SUMMARY – ALL ENVS (top movers by service)")
    print("=" * 65)
    dfs_all = build_dataframes(all_records)
    if not dfs_all.get("trend_service", pd.DataFrame()).empty:
        for _, row in dfs_all["trend_service"].head(10).iterrows():
            cols  = list(row.index)
            svc   = row[cols[0]]
            delta = row["delta_abs"]
            pct   = row["delta_pct"]
            arrow = "▲" if delta > 0 else "▼"
            print(f"  {arrow}  {svc:<48}  {'+' if delta>=0 else ''}{delta:>9.2f} USD  ({pct:+.1f}%)")
    else:
        print("  Not enough months to compute trend.")

    print("\nDone.")
    for p in xlsx_paths:
        print(f"  Excel  : {p}")
    print(f"  HTML   : {html_path}")


if __name__ == "__main__":
    main()
